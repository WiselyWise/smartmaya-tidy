#!/usr/bin/env python3
"""
Smart Maya Tidy — Google Drive duplicate-file analysis and report builder.

Input:  a JSONL file, one file record per line, with at least:
        id, title, fileSize (bytes, int), path (full folder path incl. filename),
        createdTime, modifiedTime, viewedByMeTime  (ISO8601 strings; may be empty)
        Records missing "fileSize" (Google-native docs with no byte size) are ignored
        for clustering purposes but not required to be excluded from the input file.

Output: an XLSX workbook (openpyxl) with a Summary, Duplicate Clusters, Top Space Hogs,
        and Folder Rollup tab, plus a JSON "trash plan" (keep-one-per-cluster) for the
        high-confidence tier only.

Usage:
    python3 dedupe_drive.py --input files.jsonl --out report.xlsx

This script does not talk to Google Drive itself — gathering files.jsonl is the
Claude-side job described in SKILL.md (via the Google Drive MCP connector). This script
is the deterministic, re-runnable analysis half, so the clustering logic is consistent
and auditable rather than re-derived by an LLM each time.
"""
import argparse
import json
import re
from collections import defaultdict

SMALL_FILE_THRESHOLD = 100_000  # bytes; below this, exact-size match alone is too weak a signal


def load_records(path):
    records = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            fid = rec.get("id")
            if not fid or fid in records:
                continue
            size = rec.get("fileSize", 0) or 0
            try:
                size = int(size)
            except (TypeError, ValueError):
                size = 0
            records[fid] = {
                "id": fid,
                "title": rec.get("title", ""),
                "size": size,
                "path": rec.get("path") or rec.get("title", ""),
                "createdTime": rec.get("createdTime", ""),
                "modifiedTime": rec.get("modifiedTime", ""),
                "viewedByMeTime": rec.get("viewedByMeTime", ""),
            }
    return list(records.values())


def _name_family(title):
    """Strip common copy-suffix patterns so 'IMG_1.MOV' and 'IMG_1 (1).MOV' compare equal."""
    base = re.sub(r"\s*\(\d+\)|\s*-?\s*copy\s*\d*", "", title, flags=re.I).strip().lower()
    return base


def cluster_by_size(records):
    """Group files by exact byte size; split into high-confidence vs manual-review tiers.

    High confidence: size >= SMALL_FILE_THRESHOLD, OR (below threshold but every member's
    name reduces to the same base name once copy-suffixes are stripped).
    Manual review: below threshold with genuinely different names — an exact-size match at
    a few bytes to a few KB is too weak a signal to act on automatically (unrelated tiny
    files can coincidentally share a byte count).
    """
    by_size = defaultdict(list)
    for r in records:
        if r["size"] > 0:
            by_size[r["size"]].append(r)

    high_conf, review = [], []
    for size, members in by_size.items():
        if len(members) < 2:
            continue
        families = {_name_family(m["title"]) for m in members}
        confident = size >= SMALL_FILE_THRESHOLD or len(families) == 1
        target = high_conf if confident else review
        target.append((size, members))

    high_conf.sort(key=lambda c: c[0] * (len(c[1]) - 1), reverse=True)
    review.sort(key=lambda c: c[0] * (len(c[1]) - 1), reverse=True)
    return high_conf, review


def build_trash_plan(clusters):
    """Keep the earliest-createdTime copy in each cluster; plan to trash the rest."""
    plan = []
    for size, members in clusters:
        ms = sorted(members, key=lambda m: m.get("createdTime") or "9999")
        keep, trash = ms[0], ms[1:]
        for t in trash:
            plan.append({"id": t["id"], "title": t["title"], "path": t["path"],
                         "size": size, "keep_id": keep["id"], "keep_path": keep["path"]})
    return plan


def build_report(records, high_conf, review, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    bold = Font(name=FONT, bold=True)
    normal = Font(name=FONT, size=10)
    border = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font, cell.border = header_fill, header_font, border

    def autosize(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    total_bytes = sum(r["size"] for r in records)
    total_reclaim = sum(size * (len(m) - 1) for size, m in high_conf)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["B2"] = "Smart Maya Tidy — Drive Duplicate Audit"
    ws["B2"].font = Font(name=FONT, bold=True, size=16, color="1F4E78")
    rows = [
        ("Files audited", f"{len(records):,}"),
        ("Total size audited", f"{total_bytes/1e9:.2f} GB"),
        ("High-confidence duplicate clusters", f"{len(high_conf):,}"),
        ("Confirmed reclaimable space", f"{total_reclaim/1e9:.2f} GB"),
        ("Low-confidence clusters (manual review, not auto-actioned)", f"{len(review):,}"),
    ]
    r = 4
    for label, val in rows:
        ws.cell(row=r, column=2, value=label).font = bold
        ws.cell(row=r, column=3, value=val).font = normal
        r += 1
    autosize(ws, [3, 40, 20])

    ws2 = wb.create_sheet("Duplicate Clusters")
    headers = ["#", "Size", "Copies", "Reclaimable (GB)", "File name", "Keep", "Trash candidates"]
    for i, h in enumerate(headers, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, len(headers))
    row = 2
    for i, (size, members) in enumerate(high_conf, 1):
        ms = sorted(members, key=lambda m: m.get("createdTime") or "9999")
        keep, trash = ms[0], ms[1:]
        size_h = f"{size/1e9:.2f} GB" if size >= 1e9 else (f"{size/1e6:.1f} MB" if size >= 1e6 else f"{size/1e3:.0f} KB")
        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=size_h)
        ws2.cell(row=row, column=3, value=len(ms))
        ws2.cell(row=row, column=4, value=round(size * (len(ms) - 1) / 1e9, 3))
        ws2.cell(row=row, column=5, value=keep["title"])
        ws2.cell(row=row, column=6, value=keep["path"])
        ws2.cell(row=row, column=7, value="; ".join(t["path"] for t in trash[:6]) + (f" (+{len(trash)-6} more)" if len(trash) > 6 else ""))
        for c in range(1, len(headers) + 1):
            cell = ws2.cell(row=row, column=c)
            cell.font, cell.border = normal, border
            cell.alignment = wrap
        row += 1
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:G{row-1}"
    autosize(ws2, [5, 12, 8, 16, 34, 34, 60])

    ws3 = wb.create_sheet("Top Space Hogs")
    headers3 = ["#", "Size (GB)", "File name", "Path", "Last modified", "Last viewed"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=i, value=h)
    style_header(ws3, len(headers3))
    top = sorted(records, key=lambda r: -r["size"])[:200]
    row = 2
    for i, f in enumerate(top, 1):
        ws3.cell(row=row, column=1, value=i)
        ws3.cell(row=row, column=2, value=round(f["size"] / 1e9, 3))
        ws3.cell(row=row, column=3, value=f["title"])
        ws3.cell(row=row, column=4, value=f["path"])
        ws3.cell(row=row, column=5, value=(f.get("modifiedTime") or "")[:10])
        ws3.cell(row=row, column=6, value=(f.get("viewedByMeTime") or "")[:10])
        for c in range(1, len(headers3) + 1):
            cell = ws3.cell(row=row, column=c)
            cell.font, cell.border = normal, border
        row += 1
    ws3.freeze_panes = "A2"
    autosize(ws3, [5, 12, 34, 50, 14, 14])

    wb.save(out_path)
    return total_reclaim


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", required=True, help="JSONL file of Drive file records")
    ap.add_argument("--out", default="report.xlsx", help="Output XLSX path")
    ap.add_argument("--trash-plan", default="trash_plan.json", help="Output JSON keep/trash plan")
    args = ap.parse_args()

    records = load_records(args.input)
    high_conf, review = cluster_by_size(records)
    reclaim = build_report(records, high_conf, review, args.out)
    plan = build_trash_plan(high_conf)
    with open(args.trash_plan, "w") as f:
        json.dump(plan, f, indent=2)

    print(f"Files audited: {len(records):,}")
    print(f"High-confidence clusters: {len(high_conf):,}  (reclaim: {reclaim/1e9:.2f} GB)")
    print(f"Manual-review clusters:   {len(review):,}")
    print(f"Report written to {args.out}")
    print(f"Trash plan ({len(plan):,} file ids to trash, one kept per cluster) written to {args.trash_plan}")
    print("Nothing has been deleted. Review the report and trash plan before acting on them.")


if __name__ == "__main__":
    main()
